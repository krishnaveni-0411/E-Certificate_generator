import subprocess   # Used to run external commands/programs from Python
import sys          # Provides access to Python runtime (e.g., current interpreter path)
import os           # Interacts with the operating system (files, paths, directories)

def install(package):
    # Install a package using pip via the current Python interpreter
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

# Run these installations ONLY on the first run
# install('img2pdf')    # For converting images to PDF
# install('openpyxl')   # For reading/writing Excel files (.xlsx, .xlsm, etc.)
# install('Pillow')     # For image editing and manipulation

from img2pdf import convert, AlphaChannelError   # For converting images to PDF (handles transparency issues)
from PIL import Image, ImageDraw, ImageFont     # For image editing (drawing text, fonts, templates)
from tkinter.filedialog import askopenfile      # Opens a file picker dialog to select Excel file
from openpyxl import load_workbook              # For reading/writing Excel files (.xlsx, .xlsm, etc.)

# Ask user to select an Excel workbook
file = askopenfile(
    title='Select the Workbook', 
    mode='r', 
    filetypes=[('Microsoft Excel', '.xlsx .xlsm .xltx .xltm')]
)

if file is not None:
    dirpath = os.path.dirname(file.name)   # Directory where the Excel file is located
else:
    sys.exit()                             # Exit program if no file is selected
filepath = file.name

# Paths for storing certificates
allCertPath = dirpath + "/All_Certificates"
allCertImgPath = allCertPath + "/Images"
allCertPdfPath = allCertPath + "/PDFs"

# Create folders if not already present
try:
    os.mkdir(allCertPath)
    os.mkdir(allCertImgPath)
    os.mkdir(allCertPdfPath)
except:
    {}  # Ignore if folders already exist

wb = load_workbook(filepath, data_only=True)   # Open the workbook
for ws in wb:                                  # Loop through all sheets
    for r in range(3, 5):  
    # Change these values based on your Excel file
    # Example:Here range(3, 5) → processes rows 3 to 4
        cell = ws.cell(row=r, column=1)
        if cell.value is None:                 # Stop if empty row
            break

        # Extract participant details from Excel
        # ⚠️ Update column numbers as per your Excel file structure:
        #   column=1 → ID (e.g., Roll Number / Reg No.)
        #   column=2 → Name
        #   column=3 → Department

        id = str(ws.cell(row=r, column=1).value).strip().upper()
        name = str(ws.cell(row=r, column=2).value).strip().title()
        dept = str(ws.cell(row=r, column=3).value).strip().upper()

        # Paths for output files
        eachmemberIMGpath = allCertImgPath + '/' + id + '.png'
        eachmemberPDFpath = allCertPdfPath + '/' + id + '.pdf'

        # Open certificate template
        certificate = Image.open('data_alchemy.png')
        draw = ImageDraw.Draw(certificate)

        # Set font style and size
        name_font = ImageFont.truetype('LucidaUnicodeCalligraphy.ttf', 100)

        # Add participant details to certificate
        # top = 550
        # # ⚠️ Adjust the (x, y) coordinates below according to your certificate template.
        # # You may need to tweak these values until the text is placed correctly.
        # # (900, top+1430)   → Position of Name
        # # (4300, top+1430)  → Position of ID
        # # (2340, top+1655)  → Position of Department
        # draw.text((900, top+1430), name, fill=(0, 0, 0, 255), font=name_font)
        # draw.text((4300, top+1430), id, fill=(0, 0, 0, 255), font=name_font)
        # draw.text((2340, top+1655), dept, fill=(0, 0, 0, 255), font=name_font)


        # Function to draw centered text (compatible with newer Pillow versions)
        def draw_centered_text(draw, text, font, y, fill=(0, 0, 0, 255)):
            # Get text bounding box (left, top, right, bottom)
            bbox = draw.textbbox((0, 0), text, font=font)
            text_width = bbox[2] - bbox[0]
            # Calculate centered X position
            x = (certificate.width - text_width) // 2
            # Draw text at (x, y)
            draw.text((x, y), text, fill=fill, font=font)

        # Example usage
        top = 1430   # Y position of name
        draw_centered_text(draw, name, name_font, top)

        # Save as image
        certificate = certificate.convert('RGB')
        certificate.save(eachmemberIMGpath)

        # Convert image to PDF
        certificate = Image.open(eachmemberIMGpath)
        pdf_bytes = convert(certificate.filename)
        f = open(eachmemberPDFpath, "wb")
        f.write(pdf_bytes)
        certificate.close()
        f.close()

        # Log success
        print(r, name)

# Save workbook after updates
wb.save(file.name)
print("You're Welcome!!!")   # Final message after successful execution
